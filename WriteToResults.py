import openpyxl as op
import getpass

def WriteToExcel(scores_dict, column_number):
    """
    The way this function works is that it'll be given a dictionary object to work with.

    The keys here will ALWAYS be the file's name (such as 'wrk_sl_2001.txt'), and the value will be
    whatever it is we want to write to the file. This value could be a paragraph, or a score.

    We just need to specify which column to write to.
    """
    path = '/Users/alfredosuarez/Desktop/Files/Sentiment-Analysis/Results.xlsx'
    wb_obj = op.load_workbook(path)
    sheet_obj = wb_obj.active

    current_row = 2

    if column_number == 2:
        # this would mean we're trying to write the paragraphs, so we'll also allow
        # the user to write the file names
        for key in scores_dict:
            name_cell = sheet_obj.cell(column=1, row=current_row)
            paragraph_cell = sheet_obj.cell(column=column_number, row=current_row)

            name_cell.value = key
            paragraph_cell.value = scores_dict[key]

            current_row += 1

    else:
        pass

    wb_obj.save('/Users/'+getpass.getuser()+'/Desktop/Files/Sentiment-Analysis/Results.xlsx')
